# -*- coding: utf-8 -*-
"""
registry.py — cadastro de carteiras (TemplateCarteiras.xlsx) → registry validado.
===================================================================================
Extraído de build_snapshot.py na refatoração de 2026-07-20 (CLAUDE.md §3/§4 —
separar "buscar dado" (Excel) + "validar contra o Mongo" em funções isoladas,
fora do orquestrador). Segue PLANNING.md §Fontes de Cadastro (planilhas):
colunas A-Q, 792 linhas, validadas 1:1 contra `wallets`/`entities`/`groupings`.

Nenhuma escrita: só leitura do Excel (openpyxl) e leitura das coleções
pequenas do Mongo (recebidas já carregadas via db.py, nunca conecta direto).
"""

import openpyxl

# Modelos de carga considerados "sistêmicos" (o resto = carga manual/humana) —
# PLANNING §Fontes de Cadastro, coluna K "Modelo de Carga".
MODELOS_DE_CARGA_SISTEMICOS = {"API", "AWS", "Scrapping Site", "XML"}

_COLUNAS_TEMPLATE = [
    "nome", "walletId", "instituicao", "excecao", "devePublicar", "cargaRetroativa",
    "agrupamentos", "periodicidade", "defasagem", "repeticaoDiaria", "modeloCarga",
    "precificacao", "c3Todos", "captura", "duRecebimentoPdf", "duUploadBeehus", "explosao",
]


def ler_linhas_do_template(caminho_xlsx):
    """Contexto:
    Lê `TemplateCarteiras.xlsx` (aba única, colunas A-Q, linha 1 = header) e
    devolve a lista crua de linhas do cadastro — PLANNING.md §Fontes de
    Cadastro, "Colunas confirmadas (A-Q)". Não valida nada contra o Mongo
    ainda (isso é `montar_registry_validado`, abaixo).

    [CORRIGIDO 2026-07-24, pedido do usuário] O workbook nunca era fechado
    explicitamente — o openpyxl recomenda `close()` sempre que se abre um
    arquivo, senão o handle do arquivo pode ficar preso mais tempo do que o
    necessário. Isso, somado a VÁRIOS processos `app.py` acumulados de
    reinícios anteriores (cada um tinha lido o Excel no boot e nunca
    fechado), impedia o Excel de salvar o TemplateCarteiras.xlsx enquanto o
    usuário editava. [Tentativa descartada: `read_only=True` também fecha
    direitinho, mas os testes mostraram que fica MUITO lento com o padrão de
    acesso célula-a-célula usado aqui (`planilha.cell(linha, coluna)`
    repetido) — read_only é otimizado pra iteração sequencial
    (`iter_rows()`), não pra acesso aleatório; ficou pra trás por enquanto,
    reescrever o loop é um passo maior que o pedido de hoje.]

    Pseudocódigo:
      1. Abre o workbook (data_only=True — valores calculados, nunca
         fórmula) dentro de um try/finally, garantindo o fechamento mesmo
         se algo abaixo lançar exceção.
      2. Para cada linha a partir da 2ª, lê as 17 colunas na ordem fixa.
      3. Pula linhas sem WalletID (cauda vazia da planilha).
      4. Retorna lista de dicts com as chaves de _COLUNAS_TEMPLATE.
    """
    workbook = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    try:
        planilha = workbook[workbook.sheetnames[0]]
        linhas = []
        for numero_linha in range(2, planilha.max_row + 1):
            valores = [planilha.cell(numero_linha, coluna).value
                       for coluna in range(1, len(_COLUNAS_TEMPLATE) + 1)]
            if valores[1] is None:  # sem WalletID -> linha vazia de cauda
                continue
            linhas.append(dict(zip(_COLUNAS_TEMPLATE, valores)))
        return linhas
    finally:
        workbook.close()


def _interpretar_defasagem(defasagem):
    """Contexto: converte o texto da coluna "Defasagem" do Excel ("D-3", "M",
    vazio) para o número de dias úteis usado pelo resto do código. Chamada
    por montar_registry_validado() por linha. Retorna int ou None (None =
    regime mensal, que usa outro cálculo de prazo — ver
    utils/datas.py::prazo_regime_mensal).

    Pseudocódigo:
      1. Vazio ou "M" -> None (regime mensal).
      2. Prefixo "D-" -> extrai e converte o número; falha de parse -> None.
      3. Qualquer outro texto -> None (defensivo).
    """
    if not defasagem or str(defasagem).strip().upper() == "M":
        return None
    texto = str(defasagem).strip().upper()
    if texto.startswith("D-"):
        try:
            return int(texto[2:])
        except ValueError:
            return None
    return None


def _interpretar_agrupamentos(valor_bruto):
    """Contexto: converte o texto da coluna "Agrupamentos Indexados" do
    Excel ("id1;id2;id3", "Não" ou vazio) numa lista de IDs. Chamada por
    montar_registry_validado() por linha. Retorna lista de strings (vazia
    quando não há agrupamento).

    Pseudocódigo:
      1. Vazio -> lista vazia.
      2. Texto "não"/"nao" (case-insensitive) -> lista vazia.
      3. Caso contrário, separa por ";" e remove espaços/itens vazios.
    """
    if not valor_bruto:
        return []
    texto = str(valor_bruto).strip()
    if texto.casefold() in ("não", "nao", ""):
        return []
    return [g.strip() for g in texto.split(";") if g.strip()]


def montar_registry_validado(linhas, carteiras_por_id, entidades_por_nome_ci,
                              empresas_por_id, agrupamentos_por_id):
    """Contexto:
    Valida cada linha do Excel contra `wallets`/`entities`/`groupings` (já
    carregados da API Beehus por db.py — migrado do Mongo direto em
    2026-08-05) e monta o registry final — a lista de carteiras "cadastradas"
    que o resto do app usa (nunca todas as carteiras que o token vê, só as do
    Template). Retorna (registry, orfas): `orfas` = linhas cujo WalletID não
    existe em `wallets` (não deveria acontecer, mas o cadastro pode ter
    mudado desde a última validação).

    Pseudocódigo:
      1. Para cada linha, resolve o WalletID contra `carteiras_por_id`; sem
         match -> vai para `orfas` e pula.
      2. Resolve Instituição -> entityId por nome exato (case-insensitive).
      3. Converte periodicidade/defasagem/agrupamentos pros formatos internos.
      4. Marca `isManualLoad` = Modelo de Carga fora dos sistêmicos conhecidos.
      5. Anexa metadados vindos do próprio `wallets` (startDateConsolidation,
         accountCode) — não vêm do Excel.
    """
    registry, orfas = [], []
    for linha in linhas:
        wallet_id = str(linha["walletId"]).strip()
        doc_carteira = carteiras_por_id.get(wallet_id)
        if doc_carteira is None:
            orfas.append(linha)
            continue
        nome_instituicao = (linha["instituicao"] or "").strip()
        entity_id = entidades_por_nome_ci.get(nome_instituicao.casefold())
        company_id = str(doc_carteira.get("companyId") or "")
        periodicidade = "M" if (linha["periodicidade"] or "").strip().upper() == "M" else "D"
        defasagem_du = _interpretar_defasagem(linha["defasagem"])
        modelo_carga = (linha["modeloCarga"] or "").strip()
        grouping_ids = [g for g in _interpretar_agrupamentos(linha["agrupamentos"])
                         if g in agrupamentos_por_id]
        registry.append({
            "walletId": wallet_id,
            "name": linha["nome"],
            "institution": nome_instituicao,
            "entityId": entity_id,
            "companyId": company_id,
            "company": empresas_por_id.get(company_id, {}).get("name", company_id or "—"),
            "mustPublish": str(linha["devePublicar"] or "").strip().casefold() == "sim",
            "groupingIds": grouping_ids,
            "periodicity": periodicidade,
            "lagBizDays": defasagem_du,
            "dailyRepetition": str(linha["repeticaoDiaria"] or "").strip().casefold() == "sim",
            "loadModel": modelo_carga,
            "isManualLoad": modelo_carga not in MODELOS_DE_CARGA_SISTEMICOS,
            "slaPdfReceiptDu": int(linha["duRecebimentoPdf"]) if linha["duRecebimentoPdf"] not in (None, "") else None,
            "slaUploadDu": int(linha["duUploadBeehus"]) if linha["duUploadBeehus"] not in (None, "") else None,
            "exception": linha["excecao"],
            "explosion": linha["explosao"],
            # explodedAssets — NADA a ver com "explosion" acima (aquele é
            # texto livre manual da coluna Excel "Explosão"; este é a lista
            # real de ativos que a carteira explode, vinda do Mongo —
            # wallets.securitiesForExplosion resolvido pra beehusName em
            # db.py::_resolver_nomes_explosao()) [2026-07-31, pedido do
            # usuário: "precisamos coletar a explosão de Ativos de cada
            # carteira... essa informações estão no MongoDB"].
            "explodedAssets": doc_carteira.get("explodedAssetNames") or [],
            # explodedWalletIds — [NOVO 2026-08-13, pedido do usuário: "se a
            # carteira da lista for comprada por alguma carteira da lista..."]
            # o walletId real (não só o nome) de cada ativo de explosão, só
            # quando a security tem correspondingWallet (db.py::_resolver_
            # explosao). Usado por snapshot_builder.mapear_carteiras_compradas
            # pra cruzar contra o próprio registry e achar carteiras do
            # Template compradas por outras carteiras do Template.
            "explodedWalletIds": doc_carteira.get("explodedWalletIds") or [],
            "startDateConsolidation": doc_carteira.get("startDateConsolidation"),
            "accountCode": doc_carteira.get("accountCode"),
        })
    return registry, orfas
