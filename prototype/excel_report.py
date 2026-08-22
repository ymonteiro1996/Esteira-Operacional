# -*- coding: utf-8 -*-
"""
excel_report.py — formata o snapshot já calculado como planilha .xlsx.
=========================================================================
Camada de "formatar pro consumidor" (aqui, um arquivo Excel em vez do
front-end) — separada de snapshot_builder.py (regra de negócio) e de
db.py (acesso a dado), CLAUDE.md §3. Extraído de build_snapshot.py na
refatoração de 2026-07-20; mesma saída de antes (abas Matriz + Detalhe +
ControleUpload), só reorganizado em módulo próprio.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from snapshot_builder import MOCKKEY_LETTER

# originalmente a mesma paleta do STATE_STYLES do front-end (PLANNING.md
# §Simbologia; fundo = SÓ estágio, rodada 7).
# [REVISADO 2026-07-28, pedido do usuário: "no excel o verde de fundo da Pro
# está mais escuro que o verde da Pub" — Pub (p/cD) e Pro-aguardando-
# publicação (wc) tiveram seus tons TROCADOS entre si SÓ NO EXCEL (mesmos 3
# hex de sempre — Unp/Pro/Pub — só a atribuição mudou), pra Pub virar o
# verde mais escuro/saturado. IMPORTANTE: o usuário confirmou que a TELA já
# estava certa e não deveria mudar ("a tela deveria ser mantida o que já
# estava, estava certo. somente no excel estava errado") — por isso as CSS
# vars --state-p-bg/--state-pro-bg de static/css/controle_cargas.css NÃO
# foram tocadas; ficam com os valores originais. Manter em sincronia só com
# XML_BG/XML_FG de static/js/controle_cargas/exportar.js (o outro gerador
# de Excel), não mais com o CSS.]
_PREENCHIMENTO_XLSX = {
    "p": ("8AE6D2", "134E4A"), "cD": ("8AE6D2", "134E4A"),
    "wu": ("FDE68A", "78350F"), "wc": ("DCFCE7", "15803D"),
    "miss": ("FEE2E2", "B91C1C"),
    "wait": ("F3F4F6", "9CA3AF"), "notcov": ("F9FAFB", "D1D5DB"),
}


def _escrever_aba_matriz(pasta_trabalho, snapshot):
    """Contexto: escreve a aba "Matriz" (carteira x dia útil, célula colorida
    igual ao grid da tela — PLANNING §Simbologia), chamada por
    write_excel_report(). Não retorna nada (mexe na pasta de trabalho por
    referência).

    Pseudocódigo:
      1. Escreve o cabeçalho (Company/Carteira/Instituição + janela) com
         destaque visual. [REMOVIDO 2026-08-05, pedido do usuário — migração
         API Beehus] As colunas-resumo Última Unp/Pro/Pub saíram por
         completo — sem endpoint equivalente na API (ver db.py).
      2. Para cada carteira (já ordenada por prioridade), escreve 1 linha com
         a sigla de cada dia, colorindo a célula pelo estado do dia.
      3. Congela painéis e ajusta largura das colunas.
    """
    ws = pasta_trabalho.active
    ws.title = "Matriz"
    janela = snapshot["meta"]["window"]
    cabecalho = ["Company", "Carteira", "Instituição"] + janela
    ws.append(cabecalho)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEF0EC")

    carteiras = snapshot["wallets"]  # já vem ordenado por prioridade
    for r in carteiras:
        mapa_celulas = {c["d"]: c for c in r["cells"]}
        linha = [r["company"], r["name"], r["institution"]]
        linha += [MOCKKEY_LETTER.get(mapa_celulas.get(d, {}).get("s", "notcov"), "—") for d in janela]
        ws.append(linha)
        linha_excel = ws.max_row
        for i, d in enumerate(janela):
            chave = mapa_celulas.get(d, {}).get("s", "notcov")
            fundo, texto = _PREENCHIMENTO_XLSX.get(chave, ("F9FAFB", "D1D5DB"))
            celula = ws.cell(row=linha_excel, column=4 + i)
            celula.fill = PatternFill("solid", fgColor=fundo)
            celula.font = Font(bold=True, color=texto)
            celula.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "D2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    for i in range(len(janela)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 7


def _escrever_aba_detalhe(pasta_trabalho, snapshot):
    """Contexto: escreve a aba "Detalhe" (1 linha por carteira com os campos
    de auditoria/triagem pedidos pelo usuário), chamada por
    write_excel_report(). Não retorna nada.

    Pseudocódigo:
      1. Escreve o cabeçalho com as colunas de auditoria.
      2. Para cada carteira, varre as células da janela pra extrair a última
         divergência, se há sequência quebrada, dias com issue e pior atraso.
      3. Escreve a linha com esses agregados + metadados de cadastro,
         colorindo Company/Carteira/Instituição/Status pelo mockkey do
         estado atual (mesma paleta da aba Matriz — tier foi aposentado,
         2026-07-24).
      4. Congela painéis, liga auto-filtro e ajusta largura das colunas.
    """
    ws2 = pasta_trabalho.create_sheet("Detalhe")
    cabecalho2 = ["Company", "Carteira", "Instituição", "Modelo de Carga", "Periodicidade",
                  "Status Atual",
                  "Dias de atraso (pior célula da janela)",
                  "Δ Rent Contrib x NAV (bp, última divergência na janela)",
                  "Issues pendentes (dias com issue na janela)", "Fora de sequência na janela?",
                  "Deve Publicar", "Exceção", "Explosão"]
    ws2.append(cabecalho2)
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEF0EC")

    for r in snapshot["wallets"]:
        ultima_div_bp, tem_sequencia_quebrada, dias_com_issue, pior_atraso, status_atual = "", "Não", 0, 0, ""
        mockkey_atual = None
        for c in r["cells"]:
            tt = c.get("tt", {})
            if "div" in tt:
                ultima_div_bp = round(tt["div"]["bp"], 1)
            if tt.get("seq"):
                tem_sequencia_quebrada = "Sim"
            if "issues" in tt:
                dias_com_issue += 1
            pior_atraso = max(pior_atraso, c.get("adu", 0))
        if r["cells"]:
            mockkey_atual = r["cells"][-1]["s"]
            status_atual = MOCKKEY_LETTER.get(mockkey_atual, "—")
        ws2.append([
            r["company"], r["name"], r["institution"], r["loadModel"],
            "Mensal" if r["monthly"] else "Diário",
            status_atual,
            pior_atraso, ultima_div_bp, dias_com_issue, tem_sequencia_quebrada,
            "Sim" if r["mustPublish"] else "Não", r.get("exception") or "", r.get("explosion") or "",
        ])
        linha_excel = ws2.max_row
        fundo, _ = _PREENCHIMENTO_XLSX.get(mockkey_atual, ("FFFFFF", "000000"))
        preenchimento = PatternFill("solid", fgColor=fundo)
        # colunas coloridas: Company/Carteira/Instituição (1-3) + Status
        # Atual (6 — era 9 antes de 2026-08-05, quando as 3 colunas
        # Última Unp/Pro/Pub saíram do meio; ver cabecalho2 acima).
        for coluna in (1, 2, 3, 6):
            ws2.cell(row=linha_excel, column=coluna).fill = preenchimento

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho2))}{ws2.max_row}"
    larguras = [22, 28, 14, 16, 12, 14, 14, 14, 12, 12, 14, 12, 12, 12, 24, 16]
    for i, w in enumerate(larguras, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w


def _escrever_aba_controle_upload(pasta_trabalho, snapshot):
    """Contexto: escreve a aba "ControleUpload" (espelho da aba de
    custodiantes do dashboard — janela de 25 datas terminando na última data
    com dado), chamada por write_excel_report(). Não escreve nada se o
    snapshot não trouxer o bloco (Excel externo indisponível na geração).

    Pseudocódigo:
      1. Se não há bloco custodianUpload, sai sem criar a aba.
      2. Calcula a janela de 25 colunas terminando na última data com dado.
      3. Escreve o cabeçalho (datas) e, por custodiante, o texto bruto de
         cada célula da janela, colorindo pela classificação (ok/alert/
         neutral).
      4. Congela painéis e ajusta largura das colunas.
    """
    cu = snapshot.get("custodianUpload")
    if not cu:
        return
    preenchimento_cu = {"ok": ("DCFCE7", "15803D"), "alert": ("FEF3C7", "92400E"),
                        "neutral": ("F3F4F6", "9CA3AF")}
    ws3 = pasta_trabalho.create_sheet("ControleUpload")
    datas = cu["dates"]
    ancora = cu.get("lastDateWithData") or (datas[-1] if datas else None)
    indice_fim = max((i for i, d in enumerate(datas) if d <= ancora), default=len(datas) - 1) \
        if ancora else len(datas) - 1
    indices_janela = list(range(max(0, indice_fim - 24), indice_fim + 1))
    ws3.append(["Gestor + Custodia"] + [datas[i] for i in indices_janela])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEF0EC")
    for linha in cu["rows"]:
        valores = [linha["label"]]
        for i in indices_janela:
            celula = linha["cells"][i]
            valores.append(celula["t"] if celula else "")
        ws3.append(valores)
        linha_excel = ws3.max_row
        for j, i in enumerate(indices_janela):
            celula = linha["cells"][i]
            classe = celula["c"] if celula else "neutral"
            fundo, texto = preenchimento_cu[classe]
            xc = ws3.cell(row=linha_excel, column=2 + j)
            xc.fill = PatternFill("solid", fgColor=fundo)
            xc.font = Font(color=texto)
            xc.alignment = Alignment(horizontal="center")
    ws3.freeze_panes = "B2"
    ws3.column_dimensions["A"].width = 34
    for j in range(len(indices_janela)):
        ws3.column_dimensions[get_column_letter(2 + j)].width = 12


def write_excel_report(snapshot, caminho_saida):
    """Contexto:
    Gera o .xlsx real (abre no Excel de verdade) com as abas Matriz +
    Detalhe + ControleUpload (quando disponível). Ponto de entrada público
    deste módulo — build_snapshot.py chama isso depois de montar o snapshot.

    Pseudocódigo:
      1. Cria a pasta de trabalho e escreve cada aba (funções isoladas
         acima — uma responsabilidade por função, CLAUDE.md §3).
      2. Salva no caminho pedido.
    """
    pasta_trabalho = Workbook()
    _escrever_aba_matriz(pasta_trabalho, snapshot)
    _escrever_aba_detalhe(pasta_trabalho, snapshot)
    _escrever_aba_controle_upload(pasta_trabalho, snapshot)
    pasta_trabalho.save(caminho_saida)
    print(f"Excel (openpyxl) escrito em {caminho_saida} ({os.path.getsize(caminho_saida)/1024:.0f} KB)")
