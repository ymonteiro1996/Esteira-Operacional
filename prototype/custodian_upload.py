"""
custodian_upload.py — ingestão do ControleUpload.xlsx (aba "Controle de Cargas")
================================================================================

Lê o Excel MANTIDO MANUALMENTE por uma pessoa da operação com o status diário
de upload por custodiante ("Gestor + Custodia") e devolve o bloco
`custodianUpload` pronto para entrar no snapshot.json (ver PLANNING.md,
seção "Aba Controle de Cargas (Custodiantes)").

REGRAS INEGOCIÁVEIS deste módulo:
- O arquivo é EXTERNO ao projeto (vive em "Cliente Beehus\\ControleUpload.xlsx")
  e é de OUTRA pessoa: somente leitura, NUNCA escrever/salvar nele. Este módulo
  só chama openpyxl.load_workbook(..., data_only=True) — não existe .save().
- Leitura 100% dinâmica: o número de linhas de custodiante CRESCE com o tempo
  (a pessoa adiciona custodiantes novos); nunca hardcodear contagem de linhas
  nem de colunas — em 2026-07-18 eram 14 linhas × 352 datas, amanhã pode ser mais.

Estrutura da planilha (validada em 2026-07-18):
- Aba `Planilha1`:
    linha 1  = rótulos esparsos de "Day Off" da pessoa que mantém a planilha
               (informativo; viram nota de tooltip na coluna da data).
    linha 2  = cabeçalho: col A = "Gestor + Custodia"; col B em diante = UMA
               data POR COLUNA (dias úteis, jun/2025..dez/2026). CUIDADO: as
               datas antigas vêm como datetime do Excel, mas a partir de
               mar/2026 vêm como STRING "DD/MM/YYYY" — parse_header_date()
               trata os dois formatos e normaliza para "YYYY-MM-DD".
    linha 3+ = col A = rótulo do custodiante (ex. "Mira - BTG Onshore");
               demais colunas = status em texto livre daquele dia.
- Aba `Planilha2`: uma única coluna de datas (67 dias úteis, nov/2025..fev/2026),
  sem nenhum status associado — avaliada e considerada rascunho/apoio da pessoa
  que mantém a planilha (provável lista auxiliar para gerar datas do cabeçalho).
  NÃO é usada por este módulo (decisão documentada no PLANNING.md).
"""

import datetime as dt

import openpyxl

# Caminho do Excel externo (fora do projeto — somente leitura).
CONTROLE_UPLOAD_XLSX = (r"C:\Users\efigueira\Beehus Tecnologia Ltda"
                        r"\Beehus Tecnologia Ltda - Documentos"
                        r"\Cliente Beehus\ControleUpload.xlsx")
SHEET_NAME = "Planilha1"
HEADER_ROW = 2       # linha do cabeçalho de datas
FIRST_DATA_ROW = 3   # primeira linha de custodiante
FIRST_DATE_COL = 2   # coluna B = primeira data


# ─────────────────────────────────────────────────────────────────────────
# 1) Parse de data do cabeçalho — datetime do Excel OU string "DD/MM/YYYY"
# ─────────────────────────────────────────────────────────────────────────

def parse_header_date(value):
    """Contexto:
    Normaliza 1 célula do cabeçalho de datas (linha 2 da Planilha1) para
    string "YYYY-MM-DD". Chamada por load_controle_upload() para cada coluna
    de B em diante. Retorna a data ISO ou None para célula vazia/formato
    irreconhecível (a coluna é ignorada — nunca derrubar a ingestão por 1
    célula torta).

    Pseudocódigo:
      1. Célula vazia (None) -> None.
      2. Se já é datetime/date do Excel (datas até fev/2026), formata direto.
      3. Senão, tenta parsear como string "DD/MM/YYYY" (datas de mar/2026 em
         diante, digitadas como texto); falha de parse -> None.
    """
    if value is None:
        return None
    # caso 1: objeto de data nativo (datetime herda de date)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    # caso 2: string "DD/MM/YYYY" (tolera espaços acidentais)
    s = str(value).strip()
    try:
        return dt.datetime.strptime(s, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────────────────
# 2) Classificação da célula de status (regra confirmada com o usuário)
# ─────────────────────────────────────────────────────────────────────────

def classify_status(raw):
    """Contexto:
    Classifica 1 célula de status (texto livre digitado pela operação) em
    "ok" (verde) / "neutral" (cinza) / "alert" (amarelo). Chamada por
    load_controle_upload() célula a célula. Retorna a classe (string).

    Regra confirmada com o usuário (2026-07-18) — ver PLANNING.md:
    - começa com "P" (após strip, case-insensitive) → OK verde. Cobre P, P T,
      PT, P (Sem T), P (Conferir T), P (R), P (Repl) T, P T 7/7, sufixos
      "(2/7 faltando desde ...)" etc. — o texto bruto fica no tooltip, então
      variantes informativas como "P (Conferir T)" continuam visíveis.
    - célula vazia OU exatamente "Feriado" (case-insensitive) → neutro cinza
      (os 2 únicos casos neutros; não são alerta).
    - qualquer outro texto → alerta amarelo (ex.: "Sem posição", "Repetido",
      "Coletado" sozinho, "T", "R", "-").
    - caso especial: "Coletado P T" (variante rara) foi listada pelo usuário
      como verde apesar de não começar com "P" — tratado como verde quando o
      texto começa com "Coletado" E o restante começa com "P" ("Coletado"
      sozinho ou "Coletado(validar...)" continuam amarelos).

    Pseudocódigo:
      1. Vazio (None ou string em branco) -> neutral.
      2. Exatamente "feriado" (case-insensitive) -> neutral.
      3. Começa com "p" -> ok.
      4. Caso especial "Coletado P..." -> ok.
      5. Qualquer outro texto -> alert.
    """
    if raw is None:
        return "neutral"
    txt = str(raw).strip()
    if not txt:
        return "neutral"
    low = txt.casefold()
    if low == "feriado":
        return "neutral"
    if low.startswith("p"):
        return "ok"
    # caso especial "Coletado P T" (ver docstring)
    if low.startswith("coletado"):
        rest = low[len("coletado"):].strip()
        if rest.startswith("p"):
            return "ok"
    return "alert"


# ─────────────────────────────────────────────────────────────────────────
# 3) Ingestão dinâmica da Planilha1 → bloco `custodianUpload` do snapshot
# ─────────────────────────────────────────────────────────────────────────

def load_controle_upload(xlsx_path=CONTROLE_UPLOAD_XLSX):
    """Contexto:
    Lê a Planilha1 do ControleUpload.xlsx (SOMENTE LEITURA) e devolve o
    bloco `custodianUpload` pronto para entrar no snapshot.json. Chamada por
    build_snapshot.py (_ler_controle_upload_custodiantes), best-effort — uma
    falha aqui não derruba o resto do snapshot.

    Pseudocódigo:
      1. Abre o workbook com data_only=True (valores, nunca fórmulas) e
         read_only=True (garantia extra de que nada é gravado de volta).
      2. Varre a linha 2 e monta a lista de colunas de data (col Excel → data
         ISO), tratando datetime e string "DD/MM/YYYY"; ordena
         cronologicamente por segurança (hoje o cabeçalho já vem em ordem).
      3. Captura os rótulos de "Day Off" da linha 1 e associa cada um à data
         da sua coluna (nota informativa de tooltip).
      4. Varre da linha 3 até ws.max_row incluindo TODA linha com rótulo
         preenchido na coluna A (dinâmico — o nº de custodiantes cresce).
      5. Classifica cada célula por classify_status() e guarda também o
         texto bruto original (tooltip) — célula vazia vira None no array
         (neutro).

    Retorna dict:
    {
      "sourceFile": caminho lido, "sheet": "Planilha1", "readAt": ISO,
      "dates":      ["YYYY-MM-DD", ...]              # colunas, ordem cronológica
      "rows":       [{"label": str, "cells": [None | {"c": cls, "t": raw}]}],
                    # cells alinhado 1:1 com "dates"; None = vazio (neutro)
      "dayOffNotes": {"YYYY-MM-DD": "Day Off ...", ...},
      "lastDateWithData": última data com ALGUM status preenchido (âncora da
                          janela deslizante do front-end),
      "counts":     {"ok": n, "alert": n, "neutral": n}   # todas as células
    }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[SHEET_NAME]
        # read_only mode: materializa as linhas 1x (a planilha é pequena —
        # ~17 linhas × ~353 colunas em 2026-07-18).
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()   # fecha o handle — nunca seguramos o arquivo da pessoa aberto

    if len(grid) < FIRST_DATA_ROW:
        raise ValueError(f"{xlsx_path}: aba {SHEET_NAME} tem menos de "
                         f"{FIRST_DATA_ROW} linhas — estrutura inesperada")

    dayoff_row = grid[0]              # linha 1 (índice 0)
    header_row = grid[HEADER_ROW - 1]  # linha 2 (índice 1)

    # ── passo 2: colunas de data (índice na lista → data ISO) ────────────
    # date_cols = [(idx0_da_coluna, "YYYY-MM-DD"), ...] ordenado por data.
    date_cols = []
    for idx0 in range(FIRST_DATE_COL - 1, len(header_row)):
        iso = parse_header_date(header_row[idx0])
        if iso is not None:
            date_cols.append((idx0, iso))
    date_cols.sort(key=lambda pair: pair[1])   # defensivo; hoje já vem em ordem
    dates = [iso for _, iso in date_cols]

    # ── passo 3: notas "Day Off" da linha 1 → data da coluna ─────────────
    dayoff_notes = {}
    for idx0, iso in date_cols:
        if idx0 < len(dayoff_row) and dayoff_row[idx0] is not None:
            note = str(dayoff_row[idx0]).strip()
            if note:
                dayoff_notes[iso] = note

    # ── passos 4/5: linhas de custodiante + classificação célula a célula ─
    rows_out = []
    counts = {"ok": 0, "alert": 0, "neutral": 0}
    last_date_with_data = None
    for raw_row in grid[FIRST_DATA_ROW - 1:]:
        label = raw_row[0] if raw_row else None
        if label is None or not str(label).strip():
            # linha sem rótulo na coluna A: pula (pode ser linha vazia de
            # cauda OU um respiro no meio — não interrompe a varredura)
            continue
        cells = []
        for idx0, iso in date_cols:
            raw = raw_row[idx0] if idx0 < len(raw_row) else None
            cls = classify_status(raw)
            counts[cls] += 1
            if raw is None or not str(raw).strip():
                cells.append(None)   # vazio → neutro implícito (payload menor)
            else:
                cells.append({"c": cls, "t": str(raw).strip()})
                if last_date_with_data is None or iso > last_date_with_data:
                    last_date_with_data = iso
        rows_out.append({"label": str(label).strip(), "cells": cells})

    return {
        "sourceFile": xlsx_path,
        "sheet": SHEET_NAME,
        "readAt": dt.datetime.now().isoformat(timespec="seconds"),
        "dates": dates,
        "rows": rows_out,
        "dayOffNotes": dayoff_notes,
        "lastDateWithData": last_date_with_data,
        "counts": counts,
    }


if __name__ == "__main__":
    # teste manual isolado: python custodian_upload.py
    block = load_controle_upload()
    print(f"custodiantes: {len(block['rows'])}")
    print(f"datas:        {len(block['dates'])} ({block['dates'][0]} .. {block['dates'][-1]})")
    print(f"última data com dado: {block['lastDateWithData']}")
    print(f"contagem:     {block['counts']}")
    print(f"day off:      {block['dayOffNotes']}")
    for r in block["rows"]:
        filled = sum(1 for c in r["cells"] if c)
        print(f"  - {r['label']}: {filled} células preenchidas")
